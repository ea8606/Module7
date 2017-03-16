import openpyxl.reader.excel as py


class FileOperations():
    """Class does various operatiions on excel and text file.  """

    def open_the_files(self, excel_file, text_file):
        """

        :param excel_file: name of the excel file,(the roadmap)
        :param text_file: name of the text file with test cases, (masterscript file)
        :return: a reference to the opened excel file, (wb) and a list of test names, (mfolder_list)
        """

        wb = py.load_workbook(excel_file)

        # Open file containing all the automated test cases from the MasterScript folder

        try:
            master_script_file = open(text_file, 'r')
        except FileNotFoundError:
            print("[OOPS]: Couldn't find the file ",master_script_file)

        # Put the MasterScript files into a list then close the file
        mfolder_list = master_script_file.readlines()
        master_script_file.close()

        return [wb, mfolder_list]

    def get_clean_excel_sheet(self, sheetname, wb):
        """
        :param sheetname: either "SVNMasterScriptExport" or 'InMasterScriptFolderNotInRoadMa'
        :param wb: reference to the excel workbook
        :return: a workbook with the sheetname cleared out
        """

        if sheetname in wb.sheetnames:
            remove_sheet = wb.get_sheet_by_name(sheetname)
            wb.remove(remove_sheet)
            cleared_sheet = wb.create_sheet(sheetname)
            return cleared_sheet
        else:
            cleared_sheet = wb.create_sheet(sheetname)
            return cleared_sheet

            # Clear out the InMasterScriptFolderNotInRoadMa sheet, if it exists, otherwise create it...
            #  this is where the test cases from the MasterScript folder that are not in the roadmap will go
            # if 'InMasterScriptFolderNotInRoadMa' in wb.sheetnames:
            # remove_sheet = wb.get_sheet_by_name('InMasterScriptFolderNotInRoadMa')
            # wb.remove(remove_sheet)
            # wb.create_sheet('InMasterScriptFolderNotInRoadMa')
            # else:
            # wb.create_sheet('InMasterScriptFolderNotInRoadMa')



